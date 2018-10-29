from openpyxl import load_workbook

def levenshteinDistance(s1, s2):
    if len(s1) > len(s2):
        s1, s2 = s2, s1

    distances = range(len(s1) + 1)
    for i2, c2 in enumerate(s2):
        distances_ = [i2+1]
        for i1, c1 in enumerate(s1):
            if c1 == c2:
                distances_.append(distances[i1])
            else:
                distances_.append(1 + min((distances[i1], distances[i1 + 1], distances_[-1])))
        distances = distances_
    return distances[-1]

GISFileName = 'GIS_Store_Data.xlsx'
RGDFileName = 'Rural Grocery Database.xlsx'
wbGISTest = load_workbook(GISFileName)
wbRGDTest = load_workbook(RGDFileName)
GISSheet = wbGISTest['Sheet1']
RGDSheet = wbRGDTest['Stores']
counter = 0

for i in range(2, 192):
	RGDName = RGDSheet[('A' + str(i))].value
	RGDZip = RGDSheet[('F' + str(i))].value
	RGDLat = RGDSheet[('H' + str(i))].value
	RGDLon = RGDSheet[('I' + str(i))].value
	for j in range(2, 608):
		GISName = GISSheet[('B' + str(j))].value
		GISZip = GISSheet[('F' + str(j))].value
		levDist = levenshteinDistance(RGDName, GISName)
		if levDist < 15 and RGDZip == GISZip:
			GISLat = GISSheet[('L' + str(j))].value
			GISLon = GISSheet[('M' + str(j))].value
			if RGDLat is None:
				print('\nPossible match found between ' + RGDName + '(RGD) and ' + GISName + '(GIS)!')
				response = input('Would you like to copy the lat & long data over? (y/n): ')
				if response == 'y':
					RGDSheet[('H' + str(i))] = GISLat
					RGDSheet[('I' + str(i))] = GISLon
					counter += 1
				else: print('Possible match skipped')
wbRGDTest.save(RGDFileName)
print(str(counter) + ' rows effected!')

